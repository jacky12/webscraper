from urllib.request import Request, urlopen #Modules used to access websites with their URLs
import ssl
import csv
import os
import tldextract
import openpyxl


# Arguments: url
# Return: html content of a url
def loadURL(url):
    if url[0:4] != 'http':
        url = 'http://' + url
    context = ssl._create_unverified_context()  # bypasses SSL Certificate Verfication (proabably not a good idea, but it got more sites to work)
    req = Request(url, headers={
        'User-Agent': 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.11 (KHTML, like Gecko) Chrome/23.0.1271.64 Safari/537.11',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Charset': 'ISO-8859-1,utf-8;q=0.7,*;q=0.3',
        'Accept-Encoding': 'none',
        'Accept-Language': 'en-US,en;q=0.8',
        'Connection': 'keep-alive'})  # changing the header prevents some webscraper blocking techniques
    response = urlopen(req, context=context).read()  # opens the URL and returns data (in bytes)

    return response

# Arguments:
#   -savepath: path to the directory containing the CSV
#   -filename: name of the CSV
#   -rows: list of rows (which are also lists) to display on CSV
# Returns nothing
# Effect: create a CSV file with the given savepath, filename and rows (in array form) info
def createCSVFile(savepath, filename, rows):

    savepath = os.path.join(savepath + filename)

    with open(savepath, "w", encoding = 'utf-8') as output:
        writer = csv.writer(output, lineterminator='\n')
        for row in rows:
            writer.writerow(row)

# Arguments:
#   -csvfile: path to CSV file including the CSV file name
#   -row: a list elements that represent one row of data
# Returns nothing
# Effect: write a row of data into a CSV
def writeIntoCSV(csvfile, row):
    with open(csvfile, 'a') as f:
        writer = csv.writer(f)
        writer.writerow(row)


# Arguments: url
# Returns: the domain of the url
def getDomain(url):
    domain = tldextract.extract(url)[1]
    return domain


# Argument: search object
# Returns nothing
# Effect: downloads a PDF if the search object is referencing a PDF
def downloadFile(search, savepathtopdf):
    domain = getDomain(search.url)
    response = search.response
    pathToDirectory = savepathtopdf + search.search_term + os.sep
    if not os.path.exists(pathToDirectory):
        os.makedirs(pathToDirectory)
    filename = domain
    pathToFile = pathToDirectory + filename
    try:
        file = open(pathToFile + '.pdf', 'wb')
        file.write(response)
        file.close()
        search.downloaded = True
        print("Downloaded " + filename + ".pdf")
    except:
        search.downloaded = False


# Arguments:
#   -pathToFile: directory path to the folder that contains the CSV
#   -searchList: list of search objects to be shown on the CSV
#   -key_phrases: list of phrases to look out for to put on CSV
#   -overwrite: boolean to determine if we are overwriting the existing CSV or adding onto it
# Returns nothing
# Effect: Adds all sentences with keywords to the CSV file
def sentences(pathToFile, searchList, key_phrases, overwrite):
    filename = "sentences.csv"

    if not os.path.exists(pathToFile + filename) or overwrite:
        createCSVFile(pathToFile, filename, [])

    for search in searchList:
        if search.key_sentences is not None:
            rows = []
            sentences = search.key_sentences
            sentences = set(sentences)
            for sentence in sentences:
                rows.append((search.url, search.search_term, [phrase for phrase in key_phrases if phrase in sentence.lower()], sentence))

            for row in rows:
                writeIntoCSV(pathToFile + filename, row)

# Arguments:
#   -text: a string
# Effect: omits the trailing and leading spaces
# Returns text
def omitSpaces(text):
    text = text.lstrip()
    text = text.rstrip()
    return text


# Deprecated
def findCityList(filename):
    spreadsheet = openpyxl.load_workbook(filename)  # pulls data directly from excel file
    data = spreadsheet.get_sheet_by_name('Sheet1')  # specifies which sheet to pull data from
    cities = []
    for i in range(2, data.max_row):
        if data.cell(row = i, column = 1).value is not None:
            cities.append(data.cell(row = i, column = 1).value.lower())
    return cities


def appendCityToTerm(filename, before, after):
    spreadsheet = openpyxl.load_workbook(filename)  # pulls data directly from excel file
    data = spreadsheet.get_sheet_by_name('Sheet1')  # specifies which sheet to pull data from
    search_terms = []
    cities = findCityList(filename)
    for city in cities:
        if before is not None:
            before = before + ' '
        if after is not None:
            after = ' ' + after
        search_term = before + city + after
        search_terms.append(search_term)
    return search_terms

